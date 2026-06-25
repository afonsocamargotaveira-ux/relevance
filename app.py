import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import unicodedata
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


st.set_page_config(page_title="Relevance – Avaliação de Aderência", page_icon="🎯", layout="wide")

st.markdown("""
<style>
    html,body,[data-testid="stAppViewContainer"]{background-color:#ffffff;font-family:'Segoe UI',sans-serif;}
    [data-testid="stSidebar"]{background-color:#f0f6ff;}
    .logo-container{display:flex;align-items:center;gap:12px;padding:8px 0 20px 0;}
    .logo-icon{width:44px;height:44px;background:linear-gradient(135deg,#1a6fd4,#5ba3f5);border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:22px;}
    .logo-text{font-size:28px;font-weight:700;letter-spacing:-0.5px;color:#1a6fd4;line-height:1;}
    .logo-sub{font-size:11px;color:#6b8aad;letter-spacing:1.5px;text-transform:uppercase;margin-top:2px;}
    .divider{border-top:1.5px solid #dde9f8;margin:18px 0;}
    .section-label{font-size:12px;font-weight:600;color:#1a6fd4;text-transform:uppercase;letter-spacing:1px;margin-bottom:6px;}
    .stButton>button{background:linear-gradient(135deg,#1a6fd4,#3d8ef0);color:white;border:none;border-radius:8px;padding:10px 28px;font-weight:600;font-size:15px;width:100%;transition:opacity .2s;}
    .stButton>button:hover{opacity:.88;}
    .stTextArea textarea,.stTextInput input{border:1.5px solid #c5d9f2 !important;border-radius:8px !important;}
    .result-card{background:#f4f9ff;border:1px solid #d0e4f9;border-radius:10px;padding:18px 22px;margin-bottom:10px;}
    #MainMenu,footer{visibility:hidden;}
    [data-testid="stHeader"]{background:#ffffff;}
</style>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# MOTOR DE AVALIAÇÃO
# Regras:
#   - Busca APENAS na Ementa e na Indexação
#   - Usa SOMENTE os termos digitados pelo usuário
#   - termo*   → qualquer palavra que começa com "termo"
#   - "frase"  → a frase exata deve aparecer
#   - palavra  → a palavra exata (word boundary)
#   - Score 1  → zero hits; Score 2-5 → escala por hits absolutos
# ══════════════════════════════════════════════════════════════════════════════



st.set_page_config(page_title="Relevance – Avaliação de Aderência", page_icon="🎯", layout="wide")

st.markdown("""
<style>
    html,body,[data-testid="stAppViewContainer"]{background-color:#ffffff;font-family:'Segoe UI',sans-serif;}
    [data-testid="stSidebar"]{background-color:#f0f6ff;}
    .logo-container{display:flex;align-items:center;gap:12px;padding:8px 0 20px 0;}
    .logo-icon{width:44px;height:44px;background:linear-gradient(135deg,#1a6fd4,#5ba3f5);border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:22px;}
    .logo-text{font-size:28px;font-weight:700;letter-spacing:-0.5px;color:#1a6fd4;line-height:1;}
    .logo-sub{font-size:11px;color:#6b8aad;letter-spacing:1.5px;text-transform:uppercase;margin-top:2px;}
    .divider{border-top:1.5px solid #dde9f8;margin:18px 0;}
    .section-label{font-size:12px;font-weight:600;color:#1a6fd4;text-transform:uppercase;letter-spacing:1px;margin-bottom:6px;}
    .stButton>button{background:linear-gradient(135deg,#1a6fd4,#3d8ef0);color:white;border:none;border-radius:8px;padding:10px 28px;font-weight:600;font-size:15px;width:100%;transition:opacity .2s;}
    .stButton>button:hover{opacity:.88;}
    .stTextArea textarea,.stTextInput input{border:1.5px solid #c5d9f2 !important;border-radius:8px !important;}
    .result-card{background:#f4f9ff;border:1px solid #d0e4f9;border-radius:10px;padding:18px 22px;margin-bottom:10px;}
    #MainMenu,footer{visibility:hidden;}
    [data-testid="stHeader"]{background:#ffffff;}
</style>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# MOTOR DE AVALIAÇÃO
# Regras:
#   - Busca APENAS na Ementa e na Indexação
#   - Usa SOMENTE os termos digitados pelo usuário
#   - termo*   → qualquer palavra que começa com "termo"
#   - "frase"  → a frase exata deve aparecer
#   - palavra  → a palavra exata (word boundary)
#   - Score 1  → zero hits; Score 2-5 → escala por hits absolutos
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# INTEIRO TEOR — extração de hyperlinks e busca de texto
# ══════════════════════════════════════════════════════════════════════════════




st.set_page_config(page_title="Relevance – Avaliação de Aderência", page_icon="🎯", layout="wide")

st.markdown("""
<style>
    html,body,[data-testid="stAppViewContainer"]{background-color:#ffffff;font-family:'Segoe UI',sans-serif;}
    [data-testid="stSidebar"]{background-color:#f0f6ff;}
    .logo-container{display:flex;align-items:center;gap:12px;padding:8px 0 20px 0;}
    .logo-icon{width:44px;height:44px;background:linear-gradient(135deg,#1a6fd4,#5ba3f5);border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:22px;}
    .logo-text{font-size:28px;font-weight:700;letter-spacing:-0.5px;color:#1a6fd4;line-height:1;}
    .logo-sub{font-size:11px;color:#6b8aad;letter-spacing:1.5px;text-transform:uppercase;margin-top:2px;}
    .divider{border-top:1.5px solid #dde9f8;margin:18px 0;}
    .section-label{font-size:12px;font-weight:600;color:#1a6fd4;text-transform:uppercase;letter-spacing:1px;margin-bottom:6px;}
    .stButton>button{background:linear-gradient(135deg,#1a6fd4,#3d8ef0);color:white;border:none;border-radius:8px;padding:10px 28px;font-weight:600;font-size:15px;width:100%;transition:opacity .2s;}
    .stButton>button:hover{opacity:.88;}
    .stTextArea textarea,.stTextInput input{border:1.5px solid #c5d9f2 !important;border-radius:8px !important;}
    .result-card{background:#f4f9ff;border:1px solid #d0e4f9;border-radius:10px;padding:18px 22px;margin-bottom:10px;}
    #MainMenu,footer{visibility:hidden;}
    [data-testid="stHeader"]{background:#ffffff;}
</style>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# MOTOR DE AVALIAÇÃO
# Regras:
#   - Busca APENAS na Ementa e na Indexação
#   - Usa SOMENTE os termos digitados pelo usuário
#   - termo*   → qualquer palavra que começa com "termo"
#   - "frase"  → a frase exata deve aparecer
#   - palavra  → a palavra exata (word boundary)
#   - Score 1  → zero hits; Score 2-5 → escala por hits absolutos
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# INTEIRO TEOR — extração de hyperlinks e busca de texto
# ══════════════════════════════════════════════════════════════════════════════


def norm(t: str) -> str:
    t = t.lower().strip()
    t = unicodedata.normalize("NFKD", t)
    t = "".join(c for c in t if not unicodedata.combining(c))
    t = re.sub(r"[^a-z0-9\s]", " ", t)
    return re.sub(r"\s+", " ", t).strip()


# Palavras que NUNCA devem ser aceitas como termos de busca isolados
_NOISE = {
    "de","da","do","das","dos","em","na","no","nas","nos","que","com","para",
    "por","uma","uns","um","as","os","a","o","e","ou","se","ao","aos","sobre",
    "entre","ate","apos","nao","sim","mas","mais","seu","sua","seus","suas",
    "este","esta","isso","aqui","ali","como","quando","onde","qual","quais",
    "ser","ter","foi","sao","via","lei","art","inc","par","num","nos","pelo",
    "pela","pelos","pelas","esse","essa","esses","essas","deste","desta","isto",
    "aquele","aquela","neste","nesta","numa","tambem","ainda","apenas","alem",
    "sendo","tendo","sido","seja","sejam","serao","sera","podem","pode",
    "todas","todos","todo","toda","cada","tipo","forma","outro","outra",
    "mesmo","mesma","proprio","propria","dado","dados","caso","casos",
    "rara","raras","raro","raros","nova","novo","novas","novos",
    "geral","gerais","social","sociais","civil","civis","vez","vezes",
}
_SIGLAS_OK = {"tea","bpc","pcd","sus","eca","loas","apae","cid","onu","oms"}

def _valid(v: str) -> bool:
    if not v or v in _NOISE: return False
    if len(v) <= 2: return False
    if len(v) == 3 and v not in _SIGLAS_OK: return False
    return True

def parse_terms(raw: str) -> list:
    """
    Extrai termos de busca:
      termo*         → ('prefix', 'termo')
      "frase exata"  → ('exact', 'frase exata')
      "prefix*"      → ('prefix', 'prefix')
      palavra        → ('word', 'palavra')
    Filtra stopwords e fragmentos inválidos (erros de digitação nas aspas).
    """
    terms = []
    if not raw.strip():
        return terms

    QUOTE = r'["\u201c\u201d\u2018\u2019\u00ab\u00bb]'

    # 1. Tudo entre aspas
    for m in re.finditer(QUOTE + r'([^"\u201c\u201d\u2018\u2019\u00ab\u00bb]+)' + QUOTE, raw):
        phrase = m.group(1).strip()
        if phrase.endswith('*'):
            prefix = norm(phrase[:-1]).strip()
            if len(prefix) >= 3:
                terms.append(('prefix', prefix))
        else:
            # Normalizar mantendo TODAS as palavras (incluindo "de","da","do"...)
            # A frase exata deve bater com o texto como está — stopwords incluídas
            v = norm(phrase).strip()
            if not v:
                continue
            all_words = v.split()
            sig_words = [w for w in all_words if _valid(w)]
            if not all_words:
                continue
            if len(all_words) == 1:
                if _valid(all_words[0]):
                    terms.append(('word', all_words[0]))
            else:
                # Frase: usar versão completa com stopwords para match exato
                terms.append(('exact', v))

    # 2. Fora das aspas
    raw2 = re.sub(QUOTE + r'[^"\u201c\u201d\u2018\u2019\u00ab\u00bb]*' + QUOTE, ' ', raw)
    raw2 = re.sub(r'\b(or|and|not)\b', ' ', raw2, flags=re.IGNORECASE)
    raw2 = re.sub(r'[(),;]', ' ', raw2)

    # Juntar tokens onde o asterisco ficou separado: "deficiência *" → "deficiência*"
    tokens = raw2.split()
    merged = []
    i = 0
    while i < len(tokens):
        if i + 1 < len(tokens) and tokens[i + 1] == '*':
            merged.append(tokens[i] + '*')
            i += 2
        else:
            merged.append(tokens[i])
            i += 1

    for token in merged:
        if token.endswith('*'):
            prefix = norm(token[:-1]).strip()
            if len(prefix) >= 3:
                terms.append(('prefix', prefix))
        else:
            v = norm(token).strip()
            if _valid(v):
                terms.append(('word', v))

    seen = set(); unique = []
    for t in terms:
        if t not in seen:
            seen.add(t); unique.append(t)
    return unique

def term_match(ttype: str, tval: str, tn: str, words: set) -> bool:
    """Verifica se um termo aparece no texto normalizado.
    Para type='word', aceita automaticamente variantes morfológicas:
      - plural simples (+s, +es)
      - prefixo de 6+ chars (deficiencia → deficiencias, deficiente, etc.)
    """
    if ttype == "exact":
        return bool(re.search(r"\b" + re.escape(tval) + r"\b", tn))
    elif ttype == "prefix":
        return any(w.startswith(tval) for w in words)
    else:  # word — com variantes morfológicas automáticas
        if tval in words:
            return True
        # Plural simples
        if tval + "s" in words or tval + "es" in words:
            return True
        # Singular (remove s/es final)
        if tval.endswith("s") and len(tval) > 4 and tval[:-1] in words:
            return True
        if tval.endswith("es") and len(tval) > 5 and tval[:-2] in words:
            return True
        # Prefixo automático para palavras longas (>= 6 chars)
        # Cobre: deficiencia → deficiencias, deficiente, deficiência
        if len(tval) >= 6:
            return any(w.startswith(tval) for w in words)
        return False


def score_relevance(ementa: str, indexacao: str, theme: str, extra_raw: str) -> tuple:
    """
    Avalia a aderência buscando APENAS na Ementa e na Indexação.
    Usa somente os termos digitados pelo usuário.
    """
    # Texto de busca = ementa + indexação (não outros campos)
    texto = norm(ementa + " " + indexacao)
    words = set(texto.split())

    # Parsear termos do tema e dos extras
    theme_terms = parse_terms(theme)
    extra_terms  = parse_terms(extra_raw)

    # Remover do extra o que já está no tema
    theme_vals = {v for _, v in theme_terms}
    extra_terms = [t for t in extra_terms if t[1] not in theme_vals]

    all_terms = theme_terms + extra_terms
    if not all_terms:
        return 5, "Nenhum tema/termo definido — proposição listada sem filtro de relevância."

    # Contar hits (tema vale 2x, extras valem 1x)
    ht = sum(1 for tt, tv in theme_terms if term_match(tt, tv, texto, words))
    he = sum(1 for tt, tv in extra_terms  if term_match(tt, tv, texto, words))

    if ht == 0 and he == 0:
        score = 1
    else:
        n_theme = len(theme_terms)
        n_extra = len(extra_terms)

        # Cobertura separada para tema e extras
        cov_t = ht / n_theme if n_theme > 0 else 0.0
        cov_e = he / n_extra if n_extra > 0 else 0.0

        # Score base pelo tema; extras elevam o score
        if n_extra == 0:
            # Sem extras: score só pelo tema
            if   cov_t >= 0.70: score = 5
            elif cov_t >= 0.40: score = 4
            elif cov_t >= 0.15: score = 3
            else:               score = 2
        else:
            # Com extras: tema define base, extras elevam
            # Base pelo tema
            if   cov_t >= 0.70: base = 4
            elif cov_t >= 0.30: base = 3
            elif cov_t >  0.0:  base = 2
            else:               base = 1

            # Boost pelos extras (independente do tamanho da lista)
            # he >= 5 hits → máximo boost; he >= 2 → boost parcial
            if   he >= 5: boost = 2
            elif he >= 2: boost = 1
            elif he >= 1: boost = 1
            else:         boost = 0

            # Se não há hits no tema mas há muitos extras → base mínima 2
            if base == 1 and he >= 3:
                base = 2

            score = min(5, base + boost)

            # Qualquer termo adicional encontrado → mínimo 4
            if he >= 1 and score < 4:
                score = 4

    # Justificativa
    found_t = [tv for tt, tv in theme_terms if term_match(tt, tv, texto, words)][:5]
    found_e = [tv for tt, tv in extra_terms  if term_match(tt, tv, texto, words)][:6]
    ft = ", ".join(found_t) or "nenhum"
    fe = ", ".join(found_e) or "nenhum"

    msgs = {
        5: f"Alta aderência. Tema: {ft}. Termos adicionais: {fe}.",
        4: f"Boa aderência. Tema: {ft}. Termos adicionais: {fe}.",
        3: f"Aderência moderada. Tema: {ft}. Termos adicionais: {fe}.",
        2: f"Baixa aderência. Tema: {ft}. Termos adicionais: {fe}.",
        1: f"Sem aderência identificada. Nenhum termo do tema ou dos termos adicionais encontrado na ementa ou indexação.",
    }
    return score, msgs[score]


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════════════════
def build_output_excel(df_original: pd.DataFrame, scores: list, justifications: list) -> bytes:
    orig_headers = [str(v) if pd.notna(v) else "" for v in df_original.iloc[0]]
    new_headers  = [orig_headers[0], "Índice de aderência (1-5)", "Justificativa"] + orig_headers[1:]
    wb = Workbook(); ws = wb.active; ws.title = "Avaliação"
    hfill  = PatternFill("solid", fgColor="1A6FD4")
    hfont  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap   = Alignment(wrap_text=True, vertical="top")
    thin   = Side(style="thin", color="C5D9F2")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    sfills = {
        1: PatternFill("solid", fgColor="FDDEDE"),
        2: PatternFill("solid", fgColor="FDE8CC"),
        3: PatternFill("solid", fgColor="FEF9C3"),
        4: PatternFill("solid", fgColor="D5F5E3"),
        5: PatternFill("solid", fgColor="D6EAF8"),
    }
    for ci, v in enumerate(new_headers, 1):
        c = ws.cell(row=1, column=ci, value=v)
        c.fill=hfill; c.font=hfont; c.alignment=center; c.border=brd
    for rn, (score, justif) in enumerate(zip(scores, justifications), 2):
        orig = df_original.iloc[rn - 1]
        row_vals = (
            [orig.iloc[0] if pd.notna(orig.iloc[0]) else "", score, justif]
            + [(orig.iloc[j] if pd.notna(orig.iloc[j]) else "") for j in range(1, len(orig_headers))]
        )
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(row=rn, column=ci, value=v)
            c.border=brd; c.alignment=wrap
            if ci == 2:
                c.fill=sfills.get(score, PatternFill())
                c.alignment=center
                c.font=Font(bold=True, name="Calibri")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60
    for i in range(4, len(new_headers)+1):
        ws.column_dimensions[get_column_letter(i)].width = 30
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def score_color(s):
    return {1:"#e74c3c",2:"#e67e22",3:"#f1c40f",4:"#2ecc71",5:"#1a6fd4"}.get(s,"#999")


# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
for k, v in [("results_ready",False),("scores_list",[]),("justif_list",[]),
              ("excel_bytes",None),("df_raw",None),("data_rows",None),
]:
    if k not in st.session_state:
        st.session_state[k] = v


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
    <div class="logo-container">
      <div class="logo-icon">🎯</div>
      <div><div class="logo-text">relevance</div>
           <div class="logo-sub">Avaliação de Proposições</div></div>
    </div>""", unsafe_allow_html=True)
    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

    st.markdown('<div class="section-label">📌 Tema Principal (opcional)</div>', unsafe_allow_html=True)
    main_theme = st.text_input(
        "Tema", placeholder='Ex.: pessoa com deficiência  ou  deficien*  —  deixe em branco para listar todas',
        label_visibility="collapsed"
    )

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-label">🏷️ Termos Adicionais</div>', unsafe_allow_html=True)
    st.caption("Um termo por linha, ou separados por vírgula")
    extra_terms = st.text_area(
        "Extras",
        placeholder='cadeirante*\n"cadeira de rodas"\nbraile\nlibras\nautismo\nTEA',
        height=160, label_visibility="collapsed"
    )

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-label">🚫 Termos de Exclusão</div>', unsafe_allow_html=True)
    st.caption("Se encontrado → índice 1 automático")
    exclude_terms = st.text_area(
        "Exclusão",
        placeholder='Ex.: feirante*\n"comércio ambulante"\nprecatório',
        height=100, label_visibility="collapsed"
    )
    exclude_scope = st.radio(
        "Buscar termo de exclusão:",
        options=["Ementa e Indexação", "Somente Ementa", "Somente Indexação"],
        horizontal=True,
        label_visibility="visible",
    )

    if main_theme.strip():
        tt_p = parse_terms(main_theme)
        et_p = parse_terms(extra_terms)
        st.markdown(
            f'<div style="font-size:10px;color:#6b8aad;margin-top:4px;">'
            f'🔍 <b>{len(tt_p)}</b> termo(s) no tema · <b>{len(et_p)}</b> termos adicionais</div>',
            unsafe_allow_html=True
        )

    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
    st.markdown("""
    <div style="font-size:10px;color:#8fa8c8;line-height:1.7;">
    <b>Sintaxe:</b><br>
    <code>deficien*</code> → começa com <i>deficien</i><br>
    <code>"cadeira de rodas"</code> → frase exata<br>
    <code>autismo</code> → palavra exata<br><br>
    A busca é feita apenas na <b>Ementa</b> e na <b>Indexação</b>.
    </div>""", unsafe_allow_html=True)

    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
    st.markdown('<div class="section-label">📂 Planilha de Entrada</div>', unsafe_allow_html=True)
    uploaded_file = st.file_uploader("XLSX", type=["xlsx"], label_visibility="collapsed")
    if uploaded_file is None:
        st.session_state.results_ready = False

    st.markdown("<br>", unsafe_allow_html=True)
    run_btn = st.button("▶  Avaliar Proposições", use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<h2 style="color:#1a6fd4;font-weight:700;margin-bottom:4px;">Avaliação de Aderência Temática</h2>
<p style="color:#6b8aad;margin-top:0;">Analise proposições legislativas e exporte os resultados com índice e justificativa.</p>
<div class="divider"></div>""", unsafe_allow_html=True)

if not st.session_state.results_ready and not run_btn:
    for col, icon, title, desc in zip(
        st.columns(3),
        ["📁","✏️","⬇️"],
        ["1. Carregue a planilha","2. Defina o tema","3. Exporte o resultado"],
        ["Faça upload do arquivo .xlsx","Tema e termos adicionais na barra lateral","Baixe o Excel com índices e justificativas"]
    ):
        with col:
            st.markdown(f"""<div class="result-card" style="text-align:center;">
            <div style="font-size:32px;">{icon}</div>
            <div style="font-weight:600;color:#1a6fd4;margin-top:8px;">{title}</div>
            <div style="font-size:13px;color:#6b8aad;margin-top:4px;">{desc}</div>
            </div>""", unsafe_allow_html=True)

if run_btn:
    if uploaded_file is None:
        st.warning("⚠️ Faça upload de uma planilha .xlsx."); st.stop()
    # Tema e termos adicionais são opcionais — se ambos vazios, lista tudo sem filtrar

    try:
        # Ler bytes uma vez — necessário para extrair links E para pandas
        file_bytes = uploaded_file.read()
        import io as _io
        df_raw = pd.read_excel(_io.BytesIO(file_bytes), header=None, dtype=str)
    except Exception as e:
        st.error(f"Erro ao ler o arquivo: {e}"); st.stop()

    if df_raw.shape[0] < 2:
        st.error("A planilha precisa ter cabeçalho e ao menos uma linha de dados."); st.stop()

    header    = df_raw.iloc[0].tolist()
    data_rows = df_raw.iloc[1:].reset_index(drop=True)

    # Detectar colunas de Ementa e Indexação
    def find_col(header, keywords):
        for i, h in enumerate(header):
            hn = norm(str(h))
            if any(kw in hn for kw in keywords):
                return i
        return None

    col_ementa_auto = find_col(header, ["ementa"])
    col_indexa_auto = find_col(header, ["indexa"])
    col_tema_auto   = find_col(header, ["tema"])
    if col_ementa_auto is None:
        col_ementa_auto = find_col(header, ["descricao","objeto","assunto","resumo"])
    if col_indexa_auto is None:
        col_indexa_auto = find_col(header, ["keywords","palavras","tag"])

    # Permitir seleção manual caso a detecção automática erre
    header_labels = [f"{i} — {str(header[i])[:30]}" for i in range(len(header))]
    # Valores padrão — usados se o expander não for aberto
    col_ementa    = col_ementa_auto if col_ementa_auto is not None else 0
    col_indexa    = col_indexa_auto
    col_extra_text= col_tema_auto
    with st.expander("⚙️ Configurar colunas (clique para ajustar se necessário)", expanded=False):
        st.caption("O app detectou automaticamente as colunas abaixo. Corrija se necessário.")
        c1, c2 = st.columns(2)
        with c1:
            col_ementa = st.selectbox(
                "Coluna da Ementa",
                options=list(range(len(header))),
                index=col_ementa_auto if col_ementa_auto is not None else 0,
                format_func=lambda i: header_labels[i],
            )
        with c2:
            col_indexa_opts = [None] + list(range(len(header)))
            col_indexa = st.selectbox(
                "Coluna da Indexação (opcional)",
                options=col_indexa_opts,
                index=(col_indexa_opts.index(col_indexa_auto) if col_indexa_auto in col_indexa_opts else 0),
                format_func=lambda i: "— nenhuma —" if i is None else header_labels[i],
            )
        col_extra_opts = [None] + list(range(len(header)))
        col_extra_text = st.selectbox(
            "Coluna adicional de texto (ex: Tema) — opcional",
            options=col_extra_opts,
            index=(col_extra_opts.index(col_tema_auto) if col_tema_auto in col_extra_opts else 0),
            format_func=lambda i: "— nenhuma —" if i is None else header_labels[i],
        )


    if col_ementa is None:
        st.error("Selecione ao menos a coluna da Ementa."); st.stop()

    ementa_name = str(header[col_ementa])
    indexa_name = str(header[col_indexa]) if col_indexa is not None else "—"
    extra_name  = str(header[col_extra_text]) if col_extra_text is not None else "—"
    st.info(f"📋 Lendo: **{ementa_name}** · **{indexa_name}** · **{extra_name}**")

    # Termos parseados (mostra ao usuário)
    tt_show = parse_terms(main_theme)
    et_show = parse_terms(extra_terms)
    ex_show = parse_terms(exclude_terms)
    with st.expander(f"🔍 Termos interpretados — {len(tt_show)} do tema + {len(et_show)} adicionais + {len(ex_show)} de exclusão"):
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown("**Tema:**")
            for tp, tv in tt_show:
                icon = "🔤" if tp=="word" else ("🔡" if tp=="prefix" else "💬")
                st.markdown(f"  {icon} `{tv}`{'*' if tp=='prefix' else ''}")
        with col2:
            st.markdown("**Termos adicionais:**")
            for tp, tv in et_show[:30]:
                icon = "🔤" if tp=="word" else ("🔡" if tp=="prefix" else "💬")
                st.markdown(f"  {icon} `{tv}`{'*' if tp=='prefix' else ''}")
            if len(et_show) > 30:
                st.caption(f"… e mais {len(et_show)-30} termos")
        with col3:
            st.markdown("**Exclusão:**")
            for tp, tv in ex_show:
                icon = "🔤" if tp=="word" else ("🔡" if tp=="prefix" else "💬")
                st.markdown(f"  🚫 `{tv}`{'*' if tp=='prefix' else ''}")

    scores_list, justif_list = [], []
    total = len(data_rows)

    bar = st.progress(0, text="Avaliando proposições…")

    scores_list, justif_list = [], []
    for idx in range(total):
        row = data_rows.iloc[idx]
        ementa  = str(row.iloc[col_ementa])     if col_ementa     is not None else ""
        indexa  = str(row.iloc[col_indexa])     if col_indexa     is not None else ""
        extratx = str(row.iloc[col_extra_text]) if col_extra_text is not None else ""
        ementa  = "" if ementa.lower()  in ("nan","-","none") else ementa
        indexa  = "" if indexa.lower()  in ("nan","-","none") else indexa
        extratx = "" if extratx.lower() in ("nan","-","none") else extratx
        indexa_full = " ".join(filter(None, [indexa, extratx]))
        s, j = score_relevance(ementa, indexa_full, main_theme, extra_terms)
        scores_list.append(s)
        justif_list.append(j)
        bar.progress((idx + 1) / total, text=f"Avaliando {idx+1} de {total}…")
    bar.empty()


    # Aplicar termos de exclusão
    if exclude_terms.strip():
        for idx in range(total):
            if scores_list[idx] <= 1:
                continue
            row = data_rows.iloc[idx]
            ementa  = str(row.iloc[col_ementa])     if col_ementa     is not None else ""
            indexa  = str(row.iloc[col_indexa])     if col_indexa     is not None else ""
            extratx = str(row.iloc[col_extra_text]) if col_extra_text is not None else ""
            ementa  = "" if ementa.lower()  in ("nan","-","none") else ementa
            indexa  = "" if indexa.lower()  in ("nan","-","none") else indexa
            extratx = "" if extratx.lower() in ("nan","-","none") else extratx
            indexa_full = " ".join(filter(None, [indexa, extratx]))
            excl_terms_p = parse_terms(exclude_terms)
            if exclude_scope == "Somente Ementa":
                texto_excl = norm(ementa)
            elif exclude_scope == "Somente Indexação":
                texto_excl = norm(indexa_full)
            else:
                texto_excl = norm(ementa + " " + indexa_full)
            words_excl = set(texto_excl.split())
            excl_hits = [tv for tp, tv in excl_terms_p if term_match(tp, tv, texto_excl, words_excl)]
            if excl_hits:
                scores_list[idx] = 1
                justif_list[idx] = f"Excluído ({exclude_scope.lower()}) — termo encontrado: {', '.join(excl_hits[:5])}."

    st.session_state.results_ready = True
    st.session_state.scores_list   = scores_list
    st.session_state.justif_list   = justif_list
    st.session_state.df_raw        = df_raw
    st.session_state.data_rows     = data_rows
    st.session_state.excel_bytes   = build_output_excel(df_raw, scores_list, justif_list)


if st.session_state.results_ready:
    sl = st.session_state.scores_list
    jl = st.session_state.justif_list
    dr = st.session_state.data_rows

    st.markdown("### 📊 Resumo da Avaliação")
    counts = {i: sl.count(i) for i in range(1,6)}
    avg    = np.mean(sl)

    cols = st.columns(6)
    for col, sc in zip(cols[:5], range(1,6)):
        with col:
            st.markdown(f"""<div class="result-card" style="text-align:center;padding:12px;">
            <div style="font-size:22px;font-weight:700;color:{score_color(sc)};">{counts[sc]}</div>
            <div style="font-size:11px;color:#6b8aad;margin-top:2px;">Índice {sc}</div>
            </div>""", unsafe_allow_html=True)
    with cols[5]:
        st.markdown(f"""<div class="result-card" style="text-align:center;padding:12px;border-color:#1a6fd4;">
        <div style="font-size:22px;font-weight:700;color:#1a6fd4;">{avg:.1f}</div>
        <div style="font-size:11px;color:#6b8aad;margin-top:2px;">Média</div>
        </div>""", unsafe_allow_html=True)

    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
    st.markdown("### 🔍 Pré-visualização dos Resultados")

    preview_df = pd.DataFrame({
        "Proposição":    dr.iloc[:, 0].tolist(),
        "Índice":        sl,
        "Justificativa": jl,
    })

    def style_score(val):
        c={1:"#fddede",2:"#fde8cc",3:"#fef9c3",4:"#d5f5e3",5:"#d6eaf8"}
        return f"background-color:{c.get(val,'')};font-weight:bold;text-align:center;"

    st.dataframe(
        preview_df.style.map(style_score, subset=["Índice"]),
        use_container_width=True, height=380
    )

    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
    st.markdown("### ⬇️ Exportar Resultado")
    st.download_button(
        label="📥  Baixar planilha com avaliação (.xlsx)",
        data=st.session_state.excel_bytes,
        file_name="proposicoes_avaliadas.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.caption("Índice de aderência na coluna B · Justificativa na coluna C")
